from openpyxl import load_workbook, Workbook

lista_arquivos = ["CustosAutom", "PopulacaoPOA", "SuperMercado"]


print("Iniciando Robo...")
wb = Workbook()
nome_arq_final = "resultado.xlsx"


for nome_arq in lista_arquivos:
    print("Lendo arquivo %s" % nome_arq)
    #lê cada arquivo
    arquivo = load_workbook(filename="%s.xlsx" % nome_arq)
    sheet = arquivo[nome_arq]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column

    ws = wb.create_sheet(title=nome_arq)

    #Passar os dados
    for i in range(1, max_linhas + 1):
        for j in range(1, max_colunas + 1):
            c = sheet.cell(row=i, column=j)
            ws.cell(row=i, column=j).value = c.value

print("Criando arquivo final")
wb.remove(wb['Sheet'])
wb.save(filename=nome_arq_final)