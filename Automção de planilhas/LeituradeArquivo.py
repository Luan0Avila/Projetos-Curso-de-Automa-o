from openpyxl import load_workbook

wb = load_workbook(filename="nomes.xlsx") #carrega o arquivo xlsx

planilha = wb['Planilha1'] #acessa a planilha do arq

soma_idade = 0
for i in range(2,5): #loop para printar os dados da planilha em um texto
    print("%s tem %s anos." % (planilha['A%s' % i].value, planilha['B%s' % i].value))
    soma_idade += int(planilha['B%s' % i].value)

planilha['B5'] = soma_idade #fazendo formulas no excel

planilha['A7'] = "ALUNOS"
planilha.merge_cells("A7:B7") #mescla células
planilha.unmerge_cells("A7:B7") #tira a mesclagem de células

wb.save(filename="nomes.xlsx")